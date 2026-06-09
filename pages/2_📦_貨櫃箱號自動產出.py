import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

# ================= 網頁頁面初始設定 =================
st.set_page_config(page_title="貨櫃箱號產出", page_icon="📦", layout="wide")

# 🛠️ 終極北歐風裝潢加強版 (超強制 CSS 注入)
st.markdown("""
    <style>
    /* 全站北歐風淺色背景與冷灰色調文字 */
    .stApp {
        background-color: #fcfcfc !important;
        color: #2b303a !important;
    }
    /* 側邊欄改為溫潤的淺米灰 */
    [data-testid="stSidebar"] {
        background-color: #f3f4f6 !important;
        border-right: 1px solid #e5e7eb;
    }
    /* 標題與內文颜色統一 */
    h1, h2, h3, h4, h5, h6, p, label, .stMarkdown {
        color: #2b303a !important;
    }
    /* 卡片外框 */
    .upload-card {
        background-color: #ffffff;
        padding: 24px; 
        border-radius: 12px; 
        border: 1px solid #e5e7eb;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
        margin-bottom: 20px;
    }
    /* 將上傳檔案的大背景改為北歐簡約淡灰 + 細虛線外框 */
    [data-testid="stFileUploader"] {
        background-color: #fafafa !important;
        border: 1px dashed #cbd5e1 !important;
        border-radius: 8px !important;
        padding: 10px !important;
    }
    [data-testid="stFileUploader"] section {
        background-color: transparent !important;
    }
    div[data-testid="stFileUploadDropzone"] {
        background-color: #fafafa !important;
    }
    
    /* 強制將所有 Upload 小按鈕改為「常態北歐灰底黑字」 */
    button[data-testid*="stBaseButton"] {
        background-color: #e2e8f0 !important;
        color: #475569 !important;
        border: 1px solid #cbd5e1 !important;
        transition: all 0.2s ease;
    }
    button[data-testid*="stBaseButton"]:hover {
        background-color: #cbd5e1 !important;
        color: #334155 !important;
    }
    
    /* 🎨 1. 將大標題強制改為極簡深黑色 */
    .custom-main-title {
        color: #1e293b !important;  
        font-size: 1.8rem !important;
        font-weight: 700 !important;
        margin-bottom: 0.5rem;
    }
    
    /* 🎨 2. 將啟動按鈕改為北歐風冷岩灰藍色、白字 */
    .stButton > button {
        background-color: #475569 !important; 
        border-color: #475569 !important;
        color: #ffffff !important;
        font-weight: 600 !important;
        font-size: 16px !important;
        border-radius: 8px !important;
        padding: 12px 0 !important;
        transition: all 0.2s ease !important;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05) !important;
    }
    /* 按鈕滑鼠懸停效果 */
    .stButton > button:hover {
        background-color: #334155 !important; 
        border-color: #334155 !important;
        color: #ffffff !important;
    }
    /* 強制按鈕內部的文字維持純白與粗體 */
    .stButton > button p {
        color: #ffffff !important;
        font-weight: 600 !important;
    }
    
    /* 自訂精緻的中等標題字體樣式 */
    .custom-section-title {
        color: #2b5c8f !important;
        font-size: 18px !important;
        font-weight: 600 !important;
        margin-bottom: 12px !important;
        margin-top: 5px !important;
        display: flex;
        align-items: center;
    }
    </style>
""", unsafe_allow_html=True)

# ==================== 主畫面內容 ====================
st.markdown('<div class="custom-main-title">📦 貨櫃箱號自動產出系統</div>', unsafe_allow_html=True)
st.markdown("<p style='color: #555; margin-bottom: 5px;'>請上傳原始拆櫃 Excel 報表，系統會全自動依箱數進行列數倍增、產出 H 欄序號，並將 G 欄空白者全自動塗上紅底標記。</p>", unsafe_allow_html=True)
st.markdown("---")

st.markdown('<div class="upload-card">', unsafe_allow_html=True)
st.markdown('<div class="custom-section-title">📁 請上傳拆櫃明細原始檔案 (.xlsx)</div>', unsafe_allow_html=True)
ctn_file = st.file_uploader("將檔案拖放到此處", type=["xlsx"], key="uctn", label_visibility="collapsed")
st.markdown("</div>", unsafe_allow_html=True)

if ctn_file is not None:
    if st.button("🚀 啟動貨櫃箱號自動產出", type="primary", use_container_width=True):
        with st.spinner("正在執行產出中..."):
            try:
                # ─── 100% 恢復您本地原來的讀取與結構邏輯 ───
                df = pd.read_excel(ctn_file, skiprows=4, header=None)
                
                header_names = ['col_A', 'col_B', 'col_C', 'col_D', 'col_E', 'col_F', 'col_G', 'col_H', 'col_I']
                actual_col_count = len(df.columns)
                df.columns = header_names[:actual_col_count]
                for col in header_names[actual_col_count:]:
                    df[col] = None

                # 過濾雜項
                df = df[df['col_A'].notna()]
                exclude_keywords = '合計|總計|CTN|SKU|品項'
                df = df[~df['col_A'].astype(str).str.contains(exclude_keywords, case=False, na=False)]

                # 處理 G 欄（原始箱數）空白標記
                df['is_empty_g'] = df['col_G'].isna()
                df['temp_g'] = pd.to_numeric(df['col_G'], errors='coerce').fillna(1).astype(int)
                
                # 依照 temp_g（箱數）進行行數倍增展開
                df_expanded = df.loc[df.index.repeat(df['temp_g'])].copy()

                # ─── 核心修改點：對應您提供的期望圖 (image_609ea4) ───
                # H 欄結果要是 =TEXT(G2,"0") & "-1" 的形式，每列都是 [箱數]-1 且固定不遞增
                col_h_values = []
                for _, r in df_expanded.iterrows():
                    if not r['is_empty_g'] and str(r['col_G']).strip() != "":
                        try:
                            # 完美消除浮點數小數點（如 110.0 -> 110），並在尾端固定加上 "-1"
                            clean_box_num = str(int(float(r['col_G'])))
                            col_h_values.append(f"{clean_box_num}-1")
                        except:
                            raw_box_num = str(r['col_G']).strip()
                            col_h_values.append(f"{raw_box_num}-1" if raw_box_num else "")
                    else:
                        col_h_values.append("") # 如果原本 G 欄為空，則 H 欄保持空白
                
                df_expanded['col_H'] = col_h_values
                df_expanded['col_I'] = ""  # 強制將 I 欄內容清空，確保資料列中第二列以下完全無內容
                # ───────────────────────────────────────────────────

                # 先取出紅底判斷清單，再縮減欄位
                is_empty_list = df_expanded['is_empty_g'].tolist()
                
                # 只取前 9 欄資料 (A 到 I)
                output_df = df_expanded.iloc[:, :9].copy()
                
                # 賦予最終標題
                final_headers = ['商品編號', '商品名稱', '樣式', '品項條碼', '廠商批價', '叫貨數量', '箱數', '箱數', '拆櫃日期']
                output_df.columns = final_headers
                
                # ─── 使用 openpyxl 進行記憶體模式建構與美化 ───
                wb = Workbook()
                ws = wb.active
                ws.title = "拆櫃明細"

                thin_side = Side(border_style="thin", color="000000")
                full_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                center_align = Alignment(horizontal='center', vertical='center')
                ms_font = Font(name='微軟正黑體', size=11)
                red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

                # 設定第一列（表頭）格式
                ws.append(final_headers)
                for col_idx in range(1, 10):
                    header_cell = ws.cell(row=1, column=col_idx)
                    header_cell.border = full_border
                    header_cell.alignment = center_align
                    header_cell.font = Font(name='微軟正黑體', size=11, bold=True)

                # 寫入處理完後的資料列
                for row_data in output_df.fillna("").values.tolist(): 
                    ws.append(row_data)

                # 加入篩選器
                ws.auto_filter.ref = f"A1:I{ws.max_row}"

                # 處理資料列格式與紅底（依據 is_empty_list 判斷 G 欄原先是否為空）
                for row_idx, is_empty in enumerate(is_empty_list, start=2):
                    for col_idx in range(1, 10):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = full_border
                        cell.alignment = center_align
                        cell.font = ms_font
                        if is_empty:
                            cell.fill = red_fill

                # 自動調整欄位寬度（Big5 編碼字節計算長度）
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value:
                                val_str = str(cell.value)
                                byte_len = len(val_str.encode('big5')) if val_str else 0
                                if byte_len > max_length:
                                    max_length = byte_len
                        except: pass
                    ws.column_dimensions[column].width = max_length + 4

                # 輸出為二進位流供 Streamlit 網頁下載
                excel_data = BytesIO()
                wb.save(excel_data)
                excel_data.seek(0)
                
                st.success("✨ 貨櫃箱號整理&產出完畢！請點擊下方按鈕下載成果：")
                st.download_button(
                    label="📥 點此一鍵下載全新拆櫃 Excel 檔案", 
                    data=excel_data, 
                    file_name="#義烏櫃 拆櫃明細-福北路-箱數.xlsx", 
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                    use_container_width=True
                )
            except Exception as e: 
                st.error(f"❌ 錯誤: {e}")
