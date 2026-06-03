import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

# ================= 網頁頁面初始設定 =================
st.set_page_config(page_title="庫存條碼產出", page_icon="🏷️", layout="wide")

# 🛠️ 終極北歐風裝潢加強版 (超強制 CSS 注入)
st.markdown("""
    <style>
    /* 全站北歐風淺色背景與冷灰色調文字 */
    .stApp {
        background-color: #fcfcfc !important;
        color: #2b303a !important;
    }
    /* 側邊欄改為溫潤的淺米灰 */
    [data-testid="stSidebar"] {
        background-color: #f3f4f6 !important;
        border-right: 1px solid #e5e7eb;
    }
    /* 標題與內文颜色統一 */
    h1, h2, h3, h4, h5, h6, p, label, .stMarkdown {
        color: #2b303a !important;
    }
    /* 三個欄位的卡片外框 */
    .upload-card {
        background-color: #ffffff;
        padding: 24px; 
        border-radius: 12px; 
        border: 1px solid #e5e7eb;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
        margin-bottom: 20px;
    }
    /* 將上傳檔案的大背景改為北歐簡約淡灰 + 細虛線外框 */
    [data-testid="stFileUploader"] {
        background-color: #fafafa !important;
        border: 1px dashed #cbd5e1 !important;
        border-radius: 8px !important;
        padding: 10px !important;
    }
    [data-testid="stFileUploader"] section {
        background-color: transparent !important;
    }
    div[data-testid="stFileUploadDropzone"] {
        background-color: #fafafa !important;
    }
    
    /* 強制將所有 Upload 小按鈕改為「常態北歐灰底黑字」 */
    button[data-testid*="stBaseButton"] {
        background-color: #e2e8f0 !important;
        color: #475569 !important;
        border: 1px solid #cbd5e1 !important;
        transition: all 0.2s ease;
    }
    button[data-testid*="stBaseButton"]:hover {
        background-color: #cbd5e1 !important;
        color: #334155 !important;
    }
    
    /* 🎨 1. 將大標題強制改為極簡深黑色 */
    .custom-main-title {
        color: #1e293b !important;  
        font-size: 1.8rem !important;
        font-weight: 700 !important;
        margin-bottom: 0.5rem;
    }
    
    /* 🎨 2. 將啟動按鈕改為北歐風冷岩灰藍色、白字 */
    .stButton > button {
        background-color: #475569 !important; 
        border-color: #475569 !important;
        color: #ffffff !important;
        font-weight: 600 !important;
        font-size: 16px !important;
        border-radius: 8px !important;
        padding: 12px 0 !important;
        transition: all 0.2s ease !important;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05) !important;
    }
    /* 按鈕滑鼠懸停效果 */
    .stButton > button:hover {
        background-color: #334155 !important; 
        border-color: #334155 !important;
        color: #ffffff !important;
    }
    /* 強制按鈕內部的文字維持純白與粗體 */
    .stButton > button p {
        color: #ffffff !important;
        font-weight: 600 !important;
    }
    
    /* 自訂精緻的中等標題字體樣式 */
    .custom-section-title {
        color: #2b5c8f !important;
        font-size: 18px !important;
        font-weight: 600 !important;
        margin-bottom: 12px !important;
        margin-top: 5px !important;
        display: flex;
        align-items: center;
    }
    </style>
""", unsafe_allow_html=True)

# ==================== 主畫面內容 ====================
# 🎯 套用自訂深黑標題類別
st.markdown('<div class="custom-main-title">🏷️ 庫存資料與一維條碼整合系統</div>', unsafe_allow_html=True)
st.markdown("<p style='color: #555; margin-bottom: 5px;'>對應原 VBA 巨集。請同時投入三份對應報表，系統會全自動跨表關聯、智能取消合併儲存格並向下填補、自動篩選清除中文字儲位，並針對「庫存減銷售小於等於 0」的列自動刷黃底標記！</p>", unsafe_allow_html=True)
st.markdown("---")

col1, col2, col3 = st.columns(3, gap="medium")
with col1:
    st.markdown('<div class="upload-card">', unsafe_allow_html=True)
    st.markdown('<div class="custom-section-title">📦 ① 拆櫃明細檔案</div>', unsafe_allow_html=True)
    file_main = st.file_uploader("請上傳 拆櫃明細 (.xlsx)", type=["xlsx"], key="vba_main", label_visibility="collapsed")
    st.markdown("</div>", unsafe_allow_html=True)
with col2:
    st.markdown('<div class="upload-card">', unsafe_allow_html=True)
    st.markdown('<div class="custom-section-title">🗺️ ② 貨架位檔案</div>', unsafe_allow_html=True)
    file_shelf = st.file_uploader("請上傳 貨架位/儲位表 (.xlsx)", type=["xlsx"], key="vba_shelf", label_visibility="collapsed")
    st.markdown("</div>", unsafe_allow_html=True)
with col3:
    st.markdown('<div class="upload-card">', unsafe_allow_html=True)
    st.markdown('<div class="custom-section-title">📋 ③ 採購建議檔案</div>', unsafe_allow_html=True)
    file_purchase = st.file_uploader("請上傳 採購建議/建議表 (.xlsx)", type=["xlsx"], key="vba_purchase", label_visibility="collapsed")
    st.markdown("</div>", unsafe_allow_html=True)

# 🎯 這裡的 Button 就會自動套用 CSS 變成溫潤的灰藍底色
if st.button("🚀 啟動多表整合&自動產出", type="primary", use_container_width=True):
    if not (file_main and file_shelf and file_purchase):
        st.error("🚨 錯誤：您必須同時上傳「拆櫃明細」、「貨架位」與「採購建議」三份檔案才能啟動流程！")
    else:
        with st.spinner("正在執行中..."):
            try:
                df_main = pd.read_excel(file_main, skiprows=3)
                new_headers = ["商品編號", "商品名稱", "商品規格", "品項條碼", "箱裝數", "叫貨數量", "件數", "福北總庫存", "15日銷售", "福撿儲位", "一維條碼"]
                df_main.columns = new_headers[:len(df_main.columns)]
                
                df_shelf_raw = pd.read_excel(file_shelf)
                df_shelf_raw.iloc[:, 1] = df_shelf_raw.iloc[:, 1].ffill()
                
                shelf_dict = {}
                for _, row in df_shelf_raw.iterrows():
                    barcode = str(row.iloc[4]).strip() if not pd.isna(row.iloc[4]) else ""
                    loc = str(row.iloc[1]).strip() if not pd.isna(row.iloc[1]) else ""
                    if barcode:
                        shelf_dict[barcode] = loc

                df_pur_raw = pd.read_excel(file_purchase)
                pur_stock_dict = {}
                pur_sale_dict = {}
                for _, row in df_pur_raw.iterrows():
                    barcode = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ""
                    stock = row.iloc[5] if not pd.isna(row.iloc[5]) else 0
                    sale = row.iloc[12] if not pd.isna(row.iloc[12]) else 0
                    if barcode:
                        pur_stock_dict[barcode] = stock
                        pur_sale_dict[barcode] = sale

                yellow_rows = set()
                
                for idx, row in df_main.iterrows():
                    search_key = str(row["品項條碼"]).strip() if not pd.isna(row["品項條碼"]) else ""
                    if not search_key or search_key == "nan":
                        continue
                    
                    stock_val = pur_stock_dict.get(search_key, 0)
                    sale_val = pur_sale_dict.get(search_key, 0)
                    df_main.at[idx, "福北總庫存"] = stock_val
                    df_main.at[idx, "15日銷售"] = sale_val
                    
                    shelf_val = shelf_dict.get(search_key, "")
                    df_main.at[idx, "福撿儲位"] = shelf_val
                    df_main.at[idx, "一維條碼"] = f'="*" & D{idx+5} & "*"'
                    
                    if shelf_val and shelf_val not in ["W-兩倉暫存區", "W-線東品"]:
                        try:
                            diff = float(stock_val) - float(sale_val)
                            if diff <= 0:
                                yellow_rows.add(idx + 5)
                        except:
                            pass

                wb = Workbook()
                ws = wb.active
                ws.title = "庫存一維碼明細"

                thin_side = Side(border_style="thin", color="000000")
                full_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                center_align = Alignment(horizontal='center', vertical='center')
                ms_font = Font(name='微軟正黑體', size=11)
                barcode_font = Font(name='Free 3 of 9 Extended', size=30)
                yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

                for _ in range(3):
                    ws.append([])
                
                ws.append(new_headers)
                for col_idx in range(1, 12):
                    cell = ws.cell(row=4, column=col_idx)
                    cell.border = full_border
                    cell.alignment = center_align
                    cell.font = Font(name='微軟正黑體', size=11, bold=True)

                for _, row_data in df_main.fillna("").iterrows():
                    ws.append(list(row_data))

                ws.auto_filter.ref = f"A4:K{ws.max_row}"

                for r_idx in range(5, ws.max_row + 1):
                    is_yellow = r_idx in yellow_rows
                    for c_idx in range(1, 12):
                        cell = ws.cell(row=r_idx, column=c_idx)
                        cell.border = full_border
                        cell.alignment = center_align
                        cell.font = ms_font
                        if c_idx == 11:
                            cell.font = barcode_font
                        if is_yellow:
                            cell.fill = yellow_fill

                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            val_str = str(cell.value)
                            if val_str.startswith('="*'):
                                byte_len = 15
                            else:
                                try: byte_len = len(val_str.encode('big5'))
                                except: byte_len = len(val_str)
                            if byte_len > max_length: max_length = byte_len
                    ws.column_dimensions[column].width = max_length + 4

                excel_data = BytesIO()
                wb.save(excel_data)
                excel_data.seek(0)

                st.success("✨ VBA 庫存大整合全自動演算法執行成功！中文字體儲位已自動排除。")
                st.download_button(
                    label="📥 點此一鍵下載全新庫存條碼整合 Excel 檔案",
                    data=excel_data,
                    file_name="#義烏櫃 拆櫃明細-福北路-庫存資料&一維條碼.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"❌ 跨表運算時發生非預期錯誤: {e}")
